from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import OrderedDict
from datetime import date, datetime, time
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from accounts.models import Group, GroupMembership
from scenarios.models import Scenario
from schedules.models import (
    ParticipantIdentity,
    ParticipantIdentityAlias,
    SessionParticipant,
    SessionYouTubeLink,
    TRPGSession,
    normalize_participant_name,
)


YOUTUBE_URL_RE = re.compile(r'https?://(?:www\.)?(?:youtube\.com|youtu\.be)/[^\s\n\r]+')


class Command(BaseCommand):
    help = 'TRPGスケジュール表Excel/JSONから、投入前データ生成またはセッション投入を行います'

    def add_arguments(self, parser):
        source = parser.add_mutually_exclusive_group(required=True)
        source.add_argument('--sessions-csv', help='Legacy session CSV path for name-only import')
        parser.add_argument('--participants-csv', help='Legacy participants CSV path for name-only import')
        parser.add_argument('--aliases-csv', help='Legacy participant aliases CSV path')
        parser.add_argument('--allow-duplicates', action='store_true', help='Import CSV data even when duplicate legacy rows are detected')
        source.add_argument('--excel-path', help='TRPGスケジュール表2026.xlsx のパス')
        source.add_argument('--input-json', help='事前生成した投入用JSONのパス')
        parser.add_argument('--output-json', help='Excel解析結果JSONの出力先')
        parser.add_argument('--summary-md', help='Excel解析結果サマリMarkdownの出力先')
        parser.add_argument('--extract-only', action='store_true', help='Excel解析結果だけ出力し、DBへ投入しません')
        parser.add_argument('--dry-run', action='store_true', help='DBへ書き込まず投入予定件数だけ表示します')
        parser.add_argument('--include-future', action='store_true', help='基準日より未来の行も投入対象にします')
        parser.add_argument('--today', default=None, help='過去判定の基準日 YYYY-MM-DD。未指定なら実行日')
        parser.add_argument('--group-name', default='TRPGスケジュール表2026', help='投入先グループ名')
        parser.add_argument('--default-gm-username', default=None, help='KPをユーザー解決できない場合の既定GM username')

    def handle(self, *args, **options):
        today = self._parse_today(options.get('today'))

        if options.get('excel_path'):
            payload = self._extract_excel(Path(options['excel_path']), today)
            if options.get('output_json'):
                self._write_json(Path(options['output_json']), payload)
            if options.get('summary_md'):
                self._write_summary(Path(options['summary_md']), payload, today)
            if options.get('extract_only'):
                self.stdout.write(self.style.SUCCESS(
                    f"解析のみ完了: sessions={len(payload['sessions'])}, output={options.get('output_json') or '-'}"
                ))
                return
        elif options.get('input_json'):
            payload = self._read_json(Path(options['input_json']))
        else:
            payload = self._read_csv_payload(
                sessions_path=Path(options['sessions_csv']),
                participants_path=Path(options['participants_csv']) if options.get('participants_csv') else None,
                aliases_path=Path(options['aliases_csv']) if options.get('aliases_csv') else None,
            )

        import_sessions = [
            item for item in payload.get('sessions', [])
            if options.get('include_future') or item.get('is_past', True)
        ]

        if options.get('dry_run'):
            self._write_dry_run(import_sessions, payload.get('duplicates', []))
            return

        duplicates = payload.get('duplicates') or []
        if duplicates and not options.get('allow_duplicates'):
            for duplicate in duplicates:
                self.stderr.write(self.style.WARNING(duplicate))
            raise CommandError('CSV duplicate check failed. Re-run with --dry-run to inspect or --allow-duplicates to override.')

        if not import_sessions:
            self.stdout.write(self.style.WARNING('投入対象のセッションがありません'))
            return

        stats = self._import_payload(
            payload={**payload, 'sessions': import_sessions},
            group_name=options['group_name'],
            default_gm_username=options.get('default_gm_username'),
        )
        self.stdout.write(self.style.SUCCESS(
            '投入完了: '
            f"sessions created={stats['sessions_created']} updated={stats['sessions_updated']}, "
            f"participants created={stats['participants_created']} updated={stats['participants_updated']}, "
            f"youtube_links created={stats['youtube_links_created']}"
        ))

    def _parse_today(self, value: str | None) -> date:
        if value:
            return date.fromisoformat(value)
        return timezone.localdate()

    def _read_json(self, path: Path) -> dict:
        if not path.exists():
            raise CommandError(f'JSON file not found: {path}')
        with path.open('r', encoding='utf-8') as f:
            return json.load(f)

    def _read_csv_payload(
        self,
        sessions_path: Path,
        participants_path: Path | None,
        aliases_path: Path | None,
    ) -> dict:
        session_rows = self._read_csv_rows(sessions_path)
        participant_rows = self._read_csv_rows(participants_path) if participants_path else []
        alias_rows = self._read_csv_rows(aliases_path) if aliases_path else []

        alias_payload, alias_lookup = self._build_identity_alias_payload(alias_rows)
        participants_by_session: dict[str, list[dict]] = {}
        for row in participant_rows:
            legacy_session_id = self._clean_text(row.get('legacy_session_id'))
            participant_name = self._clean_text(row.get('participant_name'))
            if not legacy_session_id or not participant_name:
                continue
            identity_key = alias_lookup.get(normalize_participant_name(participant_name))
            participants_by_session.setdefault(legacy_session_id, []).append({
                'name': participant_name,
                'role': self._csv_role(row.get('role')),
                'character_name': self._clean_text(row.get('character_name')),
                'character_sheet_url': self._clean_text(row.get('character_sheet_url')),
                'identity_key': identity_key,
            })

        sessions = []
        for row in session_rows:
            legacy_session_id = self._clean_text(row.get('legacy_session_id'))
            title = self._clean_text(row.get('title'))
            gm_name = self._clean_text(row.get('gm_name'))
            participants = list(participants_by_session.get(legacy_session_id, []))
            if gm_name and not any(item['role'] == 'gm' for item in participants):
                participants.insert(0, {
                    'name': gm_name,
                    'role': 'gm',
                    'character_name': '',
                    'character_sheet_url': '',
                    'identity_key': alias_lookup.get(normalize_participant_name(gm_name)),
                })
            sessions.append({
                'source_rows': [row.get('_line_number')],
                'legacy_session_id': legacy_session_id,
                'date': self._clean_text(row.get('date')),
                'title': title,
                'scenario_title': self._clean_text(row.get('scenario_title')) or title,
                'kp': gm_name,
                'participants': participants,
                'members': [participant['name'] for participant in participants if participant['role'] != 'gm'],
                'character_names': [participant['character_name'] for participant in participants if participant['character_name']],
                'duration_minutes': self._int_or_zero(row.get('duration_minutes')),
                'visibility': self._csv_visibility(row.get('visibility')),
                'url_or_notes': [],
                'youtube_urls': [],
                'is_past': True,
                'legacy_name_only': True,
            })

        return {
            'source': str(sessions_path),
            'generated_at': timezone.now().isoformat(),
            'sessions': sessions,
            'identity_aliases': alias_payload,
            'duplicates': self._csv_duplicate_warnings(session_rows, participant_rows, alias_rows),
        }

    def _read_csv_rows(self, path: Path) -> list[dict]:
        if not path.exists():
            raise CommandError(f'CSV file not found: {path}')
        with path.open('r', encoding='utf-8-sig', newline='') as f:
            rows = []
            for index, row in enumerate(csv.DictReader(f), start=2):
                rows.append({**row, '_line_number': index})
        return rows

    def _build_identity_alias_payload(self, alias_rows: list[dict]) -> tuple[list[dict], dict[str, str]]:
        identities: OrderedDict[str, dict] = OrderedDict()
        alias_lookup = {}
        for row in alias_rows:
            key = self._clean_text(row.get('identity_key'))
            alias = self._clean_text(row.get('alias'))
            display_name = self._clean_text(row.get('display_name')) or alias or key
            if not key or not display_name:
                continue
            identity = identities.setdefault(key, {
                'identity_key': key,
                'display_name': display_name,
                'aliases': [],
            })
            if display_name and display_name not in identity['aliases']:
                identity['aliases'].append(display_name)
            if alias and alias not in identity['aliases']:
                identity['aliases'].append(alias)
            alias_lookup[normalize_participant_name(key)] = key
            alias_lookup[normalize_participant_name(display_name)] = key
            if alias:
                alias_lookup[normalize_participant_name(alias)] = key
        return list(identities.values()), alias_lookup

    def _csv_duplicate_warnings(
        self,
        session_rows: list[dict],
        participant_rows: list[dict],
        alias_rows: list[dict],
    ) -> list[str]:
        warnings = []
        self._append_duplicate_warnings(
            warnings,
            'session legacy_session_id',
            ((self._clean_text(row.get('legacy_session_id')), row.get('_line_number')) for row in session_rows),
        )
        self._append_duplicate_warnings(
            warnings,
            'participant row',
            (
                (
                    (
                        self._clean_text(row.get('legacy_session_id')),
                        normalize_participant_name(self._clean_text(row.get('participant_name'))),
                        self._csv_role(row.get('role')),
                    ),
                    row.get('_line_number'),
                )
                for row in participant_rows
            ),
        )
        self._append_duplicate_warnings(
            warnings,
            'alias row',
            (
                (
                    (
                        self._clean_text(row.get('identity_key')),
                        normalize_participant_name(self._clean_text(row.get('alias'))),
                    ),
                    row.get('_line_number'),
                )
                for row in alias_rows
            ),
        )
        return warnings

    def _append_duplicate_warnings(self, warnings: list[str], label: str, keyed_rows) -> None:
        seen = {}
        for key, line_number in keyed_rows:
            if not key or key == ('', ''):
                continue
            if key in seen:
                warnings.append(f'duplicate {label}: {key} at lines {seen[key]}, {line_number}')
            else:
                seen[key] = line_number

    def _csv_role(self, value: str | None) -> str:
        role = self._clean_text(value).lower()
        return 'gm' if role in {'gm', 'kp'} else 'player'

    def _csv_visibility(self, value: str | None) -> str:
        visibility = self._clean_text(value).lower()
        return visibility if visibility in {'private', 'group', 'link', 'public'} else 'group'

    def _int_or_zero(self, value) -> int:
        try:
            return int(value or 0)
        except (TypeError, ValueError):
            return 0

    def _write_json(self, path: Path, payload: dict) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open('w', encoding='utf-8', newline='\n') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write('\n')

    def _extract_excel(self, path: Path, today: date) -> dict:
        if not path.exists():
            raise CommandError(f'Excel file not found: {path}')

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError('Excel解析には openpyxl が必要です。AWS/本番投入時は --input-json を使用してください。') from exc

        workbook = load_workbook(path, read_only=False, data_only=True)
        if len(workbook.worksheets) < 2:
            raise CommandError('想定シート（2026 / 2026動画）が見つかりません')

        main_sheet = workbook.worksheets[0]
        video_sheet = workbook.worksheets[1]
        participants = self._extract_participants(main_sheet)
        raw_rows = self._extract_video_rows(video_sheet, today)
        sessions = self._merge_video_rows(raw_rows)
        return {
            'source': str(path),
            'generated_at': timezone.now().isoformat(),
            'today': today.isoformat(),
            'main_sheet': {
                'name': main_sheet.title,
                'participants': participants,
                'date_range': self._date_range(main_sheet),
            },
            'video_sheet': {
                'name': video_sheet.title,
                'raw_row_count': len(raw_rows),
            },
            'sessions': sessions,
        }

    def _extract_participants(self, sheet) -> list[str]:
        names = []
        for row in range(4, 80):
            value = sheet.cell(row, 1).value
            if value in (None, ''):
                break
            name = str(value).strip()
            if name and name != '累計':
                names.append(name)
        return names

    def _date_range(self, sheet) -> dict:
        dates = []
        for col in range(3, sheet.max_column + 1):
            value = sheet.cell(1, col).value
            if isinstance(value, datetime):
                dates.append(value.date())
            elif isinstance(value, date):
                dates.append(value)
        if not dates:
            return {}
        return {'start': min(dates).isoformat(), 'end': max(dates).isoformat(), 'count': len(dates)}

    def _extract_video_rows(self, sheet, today: date) -> list[dict]:
        rows = []
        for row in range(2, sheet.max_row + 1):
            values = [sheet.cell(row, col).value for col in range(1, 11)]
            if not any(value not in (None, '') for value in values):
                continue

            row_date = self._coerce_date(values[0])
            title = self._clean_text(values[2])
            url_or_note = self._clean_text(values[3])
            kp = self._clean_text(values[4])
            members = self._split_names(values[5])
            character_names = self._split_names(values[6])
            row_data = {
                'source_row': row,
                'date': row_date.isoformat() if row_date else '',
                'title': title,
                'kp': kp,
                'members': members,
                'character_names': character_names,
                'url_or_note': url_or_note,
                'is_past': bool(row_date and row_date <= today),
                'needs_review': self._row_review_flags(row_date, title, kp, members),
            }
            rows.append(row_data)
        return rows

    def _merge_video_rows(self, rows: list[dict]) -> list[dict]:
        grouped: OrderedDict[tuple[str, str], dict] = OrderedDict()
        for row in rows:
            title = row['title'] or f"未設定セッション {row['date']} row{row['source_row']}"
            key = (row['date'], title)
            if key not in grouped:
                grouped[key] = {
                    'source_rows': [],
                    'date': row['date'],
                    'title': title,
                    'kp': '',
                    'members': [],
                    'character_names': [],
                    'url_or_notes': [],
                    'youtube_urls': [],
                    'is_past': row['is_past'],
                    'needs_review': [],
                }

            session = grouped[key]
            session['source_rows'].append(row['source_row'])
            if row['kp'] and not session['kp']:
                session['kp'] = row['kp']
            self._extend_unique(session['members'], row['members'])
            self._extend_unique(session['character_names'], row['character_names'])
            if row['url_or_note']:
                self._extend_unique(session['url_or_notes'], self._split_lines(row['url_or_note']))
                self._extend_unique(session['youtube_urls'], YOUTUBE_URL_RE.findall(row['url_or_note']))
            self._extend_unique(session['needs_review'], row['needs_review'])
            if not row['title']:
                self._extend_unique(session['needs_review'], ['blank_title'])
            if len(row['source_row'] if isinstance(row['source_row'], list) else [row['source_row']]) > 1:
                self._extend_unique(session['needs_review'], ['merged_rows'])

        for session in grouped.values():
            if len(session['source_rows']) > 1:
                self._extend_unique(session['needs_review'], ['merged_rows'])
            if not session['youtube_urls']:
                self._extend_unique(session['needs_review'], ['no_youtube_url'])
            if session['kp'] == '設定なし' or '設定なし' in session['members']:
                self._extend_unique(session['needs_review'], ['unresolved_kp_or_members'])
        return list(grouped.values())

    def _row_review_flags(self, row_date, title, kp, members) -> list[str]:
        flags = []
        if not row_date:
            flags.append('invalid_date')
        if not title:
            flags.append('blank_title')
        if not kp:
            flags.append('blank_kp')
        if kp == '設定なし' or '設定なし' in members:
            flags.append('unresolved_kp_or_members')
        return flags

    def _coerce_date(self, value) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value.strip():
            try:
                return datetime.fromisoformat(value.strip()).date()
            except ValueError:
                return None
        return None

    def _clean_text(self, value) -> str:
        if value in (None, ''):
            return ''
        return str(value).strip()

    def _split_names(self, value) -> list[str]:
        text = self._clean_text(value)
        if not text:
            return []
        return [part.strip() for part in re.split(r'[\n/]+', text) if part.strip()]

    def _split_lines(self, value: str) -> list[str]:
        return [part.strip() for part in re.split(r'[\n\r]+', value) if part.strip()]

    def _extend_unique(self, target: list, values: list) -> None:
        for value in values:
            if value and value not in target:
                target.append(value)

    def _write_summary(self, path: Path, payload: dict, today: date) -> None:
        sessions = payload.get('sessions', [])
        past = [s for s in sessions if s.get('is_past')]
        future = [s for s in sessions if not s.get('is_past')]
        needs_review = [s for s in sessions if s.get('needs_review')]
        with_url = [s for s in sessions if s.get('youtube_urls')]

        lines = [
            '# TRPGスケジュール表2026 投入前確認',
            '',
            f"- Source: `{payload.get('source')}`",
            f"- Generated at: `{payload.get('generated_at')}`",
            f"- Past cutoff: `{today.isoformat()}`",
            f"- Main sheet: `{payload.get('main_sheet', {}).get('name')}`",
            f"- Video sheet: `{payload.get('video_sheet', {}).get('name')}`",
            '',
            '## Summary',
            '',
            f"- セッション候補: {len(sessions)}",
            f"- 過去対象: {len(past)}",
            f"- 未来対象: {len(future)}",
            f"- YouTube URLあり: {len(with_url)}",
            f"- 要確認: {len(needs_review)}",
            '',
            '## Usage',
            '',
            '```bash',
            'python manage.py import_trpg_schedule --input-json docs/imports/trpg_schedule_2026_pre_import.json --dry-run',
            'python manage.py import_trpg_schedule --input-json docs/imports/trpg_schedule_2026_pre_import.json --group-name "TRPGスケジュール表2026" --default-gm-username <username>',
            '```',
            '',
            '## 要確認セッション',
            '',
        ]

        for session in needs_review:
            title = self._summary_text(session.get('title'))
            lines.append(
                f"- {session.get('date')} `{title}` "
                f"rows={session.get('source_rows')} flags={', '.join(session.get('needs_review', []))}"
            )

        lines.extend(['', '## 投入候補一覧', ''])
        for session in sessions:
            title = self._summary_text(session.get('title'))
            members = ', '.join(session.get('members') or [])
            lines.append(
                f"- {session.get('date')} `{title}` KP={session.get('kp') or '-'} "
                f"PL=[{members}] rows={session.get('source_rows')}"
            )

        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text('\n'.join(lines) + '\n', encoding='utf-8')

    def _summary_text(self, value: str | None) -> str:
        return re.sub(r'\s+', ' ', value or '').strip()

    def _write_dry_run(self, sessions: list[dict], duplicates: list[str] | None = None) -> None:
        youtube_count = sum(len(item.get('youtube_urls') or self._extract_urls(item)) for item in sessions)
        participant_count = sum(
            len(item.get('participants', [])) if item.get('participants') is not None
            else len(item.get('members', [])) + (1 if item.get('kp') else 0)
            for item in sessions
        )
        self.stdout.write('DRY RUN')
        self.stdout.write(f'- sessions: {len(sessions)}')
        self.stdout.write(f'- participants including GM rows: {participant_count}')
        self.stdout.write(f'- youtube links: {youtube_count}')
        self.stdout.write(f"- duplicates: {len(duplicates or [])}")
        for duplicate in duplicates or []:
            self.stdout.write(f'  - {duplicate}')

    @transaction.atomic
    def _import_payload(self, payload: dict, group_name: str, default_gm_username: str | None) -> dict:
        stats = {
            'sessions_created': 0,
            'sessions_updated': 0,
            'participants_created': 0,
            'participants_updated': 0,
            'youtube_links_created': 0,
        }

        default_gm = self._get_default_gm(default_gm_username)
        group = self._get_or_create_group(group_name, default_gm)
        identity_lookup = self._import_identity_aliases(payload.get('identity_aliases', []))

        for item in payload.get('sessions', []):
            session_date = self._session_datetime(item.get('date'))
            if not session_date:
                continue

            gm = default_gm if item.get('legacy_name_only') else (self._resolve_user(item.get('kp')) or default_gm)
            if not gm:
                raise CommandError(f"GMを解決できません: title={item.get('title')} kp={item.get('kp')}")

            GroupMembership.objects.get_or_create(group=group, user=gm, defaults={'role': 'admin'})
            scenario = self._get_or_create_scenario(item.get('scenario_title') or item.get('title'), gm)
            description = self._build_description(payload, item)
            visibility = item.get('visibility') if item.get('visibility') in {'private', 'group', 'link', 'public'} else 'group'
            session, created = TRPGSession.objects.get_or_create(
                group=group,
                date=session_date,
                title=item.get('title') or '未設定セッション',
                defaults={
                    'gm': gm,
                    'scenario': scenario,
                    'status': 'completed',
                    'visibility': visibility,
                    'description': description,
                    'duration_minutes': item.get('duration_minutes') or 0,
                    'youtube_url': (item.get('youtube_urls') or self._extract_urls(item) or [''])[0],
                },
            )

            if created:
                stats['sessions_created'] += 1
            else:
                changed = False
                for field, value in {
                    'gm': gm,
                    'scenario': scenario,
                    'status': 'completed',
                    'visibility': visibility,
                    'description': description,
                    'duration_minutes': item.get('duration_minutes') or 0,
                }.items():
                    if getattr(session, field) != value:
                        setattr(session, field, value)
                        changed = True
                first_url = (item.get('youtube_urls') or self._extract_urls(item) or [''])[0]
                if first_url and session.youtube_url != first_url:
                    session.youtube_url = first_url
                    changed = True
                if changed:
                    session.save()
                    stats['sessions_updated'] += 1

            created_participants, updated_participants = self._ensure_participants(
                session,
                gm,
                item.get('members', []),
                item.get('character_names', []),
                participants_data=item.get('participants'),
                identity_lookup=identity_lookup,
                name_only=item.get('legacy_name_only', False),
            )
            stats['participants_created'] += created_participants
            stats['participants_updated'] += updated_participants
            stats['youtube_links_created'] += self._ensure_youtube_links(session, gm, item)

        return stats

    def _import_identity_aliases(self, aliases: list[dict]) -> dict[str, ParticipantIdentity]:
        identity_lookup = {}
        for data in aliases:
            key = self._clean_text(data.get('identity_key'))
            display_name = self._clean_text(data.get('display_name'))
            if not key or not display_name:
                continue
            identity, _ = ParticipantIdentity.objects.update_or_create(
                legacy_source='csv',
                legacy_key=key,
                defaults={'display_name': display_name},
            )
            identity_lookup[key] = identity
            for alias in data.get('aliases') or []:
                normalized_alias = normalize_participant_name(alias)
                if not normalized_alias:
                    continue
                exists = identity.aliases.filter(normalized_alias=normalized_alias, source='csv').exists()
                if not exists:
                    ParticipantIdentityAlias.objects.create(
                        identity=identity,
                        alias=alias,
                        source='csv',
                    )
        return identity_lookup

    def _get_default_gm(self, username: str | None):
        User = get_user_model()
        if username:
            user = User.objects.filter(username=username).first()
            if not user:
                raise CommandError(f'default GM user not found: {username}')
            return user
        return User.objects.filter(is_superuser=True).order_by('id').first() or User.objects.order_by('id').first()

    def _get_or_create_group(self, group_name: str, created_by):
        if not created_by:
            raise CommandError('グループ作成用ユーザーが見つかりません')
        group, _ = Group.objects.get_or_create(
            name=group_name,
            defaults={
                'created_by': created_by,
                'visibility': 'private',
                'description': 'TRPGスケジュール表2026から移行した過去セッション用グループ',
            },
        )
        GroupMembership.objects.get_or_create(group=group, user=created_by, defaults={'role': 'admin'})
        return group

    def _resolve_user(self, name: str | None):
        name = (name or '').strip()
        if not name or name == '設定なし':
            return None
        User = get_user_model()
        return (
            User.objects.filter(nickname=name).first()
            or User.objects.filter(username=name).first()
            or User.objects.filter(username__iexact=name).first()
        )

    def _get_or_create_scenario(self, title: str | None, created_by):
        scenario_title = title or '未設定シナリオ'
        scenario, _ = Scenario.objects.get_or_create(
            title=scenario_title,
            defaults={
                'created_by': created_by,
                'game_system': 'coc',
                'difficulty': 'intermediate',
                'estimated_duration': 'medium',
            },
        )
        return scenario

    def _session_datetime(self, value: str | None):
        session_date = self._coerce_date(value)
        if not session_date:
            return None
        dt = datetime.combine(session_date, time(0, 0))
        return timezone.make_aware(dt, timezone.get_current_timezone()) if timezone.is_naive(dt) else dt

    def _build_description(self, payload: dict, item: dict) -> str:
        lines = [
            'TRPGスケジュール表2026から移行',
            f"source={payload.get('source', '')}",
            f"source_rows={','.join(str(row) for row in item.get('source_rows', []))}",
        ]
        if item.get('legacy_session_id'):
            lines.append(f"legacy_session_id={item.get('legacy_session_id')}")
        notes = [note for note in item.get('url_or_notes', []) if not YOUTUBE_URL_RE.search(note)]
        if notes:
            lines.append('')
            lines.append('URL欄メモ:')
            lines.extend(f'- {note}' for note in notes)
        flags = item.get('needs_review') or []
        if flags:
            lines.append('')
            lines.append('要確認: ' + ', '.join(flags))
        return '\n'.join(lines)

    def _ensure_participants(
        self,
        session,
        gm,
        members: list[str],
        character_names: list[str],
        participants_data: list[dict] | None = None,
        identity_lookup: dict[str, ParticipantIdentity] | None = None,
        name_only: bool = False,
    ) -> tuple[int, int]:
        created_count = 0
        updated_count = 0
        if name_only:
            return self._ensure_name_only_participants(
                session,
                participants_data or [],
                identity_lookup or {},
            )

        _, created = SessionParticipant.objects.get_or_create(
            session=session,
            user=gm,
            defaults={'role': 'gm'},
        )
        if created:
            created_count += 1

        for member in members:
            if not member or member == '????' or member == (gm.nickname or gm.username):
                continue
            user = self._resolve_user(member)
            character_name = self._character_name_for_member(member, members, character_names)
            defaults = {
                'role': 'player',
                'character_name': character_name,
            }
            if user:
                GroupMembership.objects.get_or_create(group=session.group, user=user, defaults={'role': 'member'})
                participant, created = SessionParticipant.objects.get_or_create(
                    session=session,
                    user=user,
                    defaults=defaults,
                )
            else:
                participant, created = SessionParticipant.objects.get_or_create(
                    session=session,
                    user=None,
                    guest_name=member,
                    defaults=defaults,
                )
            if created:
                created_count += 1
            elif character_name and not participant.character_name:
                participant.character_name = character_name
                participant.save(update_fields=['character_name'])
                updated_count += 1
        return created_count, updated_count

    def _ensure_name_only_participants(
        self,
        session,
        participants_data: list[dict],
        identity_lookup: dict[str, ParticipantIdentity],
    ) -> tuple[int, int]:
        created_count = 0
        updated_count = 0
        for participant_data in participants_data:
            name = self._clean_text(participant_data.get('name'))
            if not name:
                continue
            role = participant_data.get('role') if participant_data.get('role') in {'gm', 'player'} else 'player'
            identity = self._identity_for_participant(participant_data, identity_lookup)
            defaults = {
                'role': role,
                'character_name': self._clean_text(participant_data.get('character_name')),
                'character_sheet_url': self._clean_text(participant_data.get('character_sheet_url')),
                'participant_identity': identity,
            }
            participant, created = SessionParticipant.objects.get_or_create(
                session=session,
                user=None,
                guest_name=name,
                defaults=defaults,
            )
            if created:
                created_count += 1
                continue

            changed_fields = []
            for field, value in defaults.items():
                if getattr(participant, field) != value:
                    setattr(participant, field, value)
                    changed_fields.append(field)
            if changed_fields:
                participant.save(update_fields=changed_fields)
                updated_count += 1
        return created_count, updated_count

    def _identity_for_participant(
        self,
        participant_data: dict,
        identity_lookup: dict[str, ParticipantIdentity],
    ) -> ParticipantIdentity:
        key = self._clean_text(participant_data.get('identity_key'))
        if key and key in identity_lookup:
            return identity_lookup[key]

        name = self._clean_text(participant_data.get('name'))
        fallback_key = f"name:{normalize_participant_name(name)}"
        identity, _ = ParticipantIdentity.objects.update_or_create(
            legacy_source='csv',
            legacy_key=fallback_key,
            defaults={'display_name': name},
        )
        identity_lookup[fallback_key] = identity
        return identity

    def _character_name_for_member(self, member: str, members: list[str], character_names: list[str]) -> str:
        if not character_names:
            return ''

        if len(members) == len(character_names):
            try:
                candidate = character_names[members.index(member)]
            except ValueError:
                candidate = ''
            if candidate:
                return candidate

        normalized_member = self._normalize_participant_name(member)
        matches = [
            name for name in character_names
            if self._normalize_participant_name(name).startswith(normalized_member)
        ]
        if matches:
            return ' / '.join(matches)
        return ''

    def _normalize_participant_name(self, value: str) -> str:
        text = unicodedata.normalize('NFKC', value or '')
        text = re.sub(r'[?(].*?[?)]', '', text)
        text = re.sub(r'\s+', '', text)
        chars = []
        for char in text:
            code = ord(char)
            if 0x30A1 <= code <= 0x30F6:
                chars.append(chr(code - 0x60))
            else:
                chars.append(char)
        return ''.join(chars).lower()

    def _ensure_youtube_links(self, session, gm, item: dict) -> int:
        created_count = 0
        urls = item.get('youtube_urls') or self._extract_urls(item)
        for order, url in enumerate(urls, start=1):
            video_id = SessionYouTubeLink.extract_video_id(url)
            if not video_id:
                continue
            _, created = SessionYouTubeLink.objects.get_or_create(
                session=session,
                video_id=video_id,
                defaults={
                    'added_by': gm,
                    'youtube_url': url,
                    'title': (item.get('title') or 'YouTube動画')[:200],
                    'duration_seconds': 0,
                    'order': order,
                },
            )
            if created:
                created_count += 1
        return created_count

    def _extract_urls(self, item: dict) -> list[str]:
        urls = []
        for value in item.get('url_or_notes', []):
            self._extend_unique(urls, YOUTUBE_URL_RE.findall(value))
        return urls
